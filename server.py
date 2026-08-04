# -*- coding: utf-8 -*-
"""
DATA VISTA '26 - Registration + Leaderboard + Email + Attendance Server
Run: python server.py
Visit: http://localhost:5000   |   Admin: http://localhost:5000/admin
"""

from flask import Flask, request, jsonify, send_from_directory, session, make_response, send_file
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json, random, string, smtplib, threading, socket, base64
import urllib.request as _urllib_req
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from urllib.parse import quote
from datetime import datetime

# ── Config ───────────────────────────────────────────
BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH       = os.path.join(BASE_DIR, 'registrations.xlsx')
LEADERBOARD_PATH = os.path.join(BASE_DIR, 'leaderboard.json')
ADMIN_USER       = os.environ.get('ADMIN_USER', 'admin')
ADMIN_PASS       = os.environ.get('ADMIN_PASS', 'DataVista@26')

HEADERS = [
    'S.No', 'Reg. ID', 'Event', 'College', 'Department',
    'Participant 1 Name', 'Participant 1 Email',
    'Participant 2 Name', 'Participant 2 Email',
    'Phone', 'Registered On',
    'Check-In Status', 'Check-In Time'
]
COL_WIDTHS = [6, 18, 20, 28, 22, 26, 32, 26, 32, 16, 22, 15, 20]

# Column indices (0-based)
IDX_SNO      = 0
IDX_REG_ID   = 1
IDX_EVENT    = 2
IDX_COLLEGE  = 3
IDX_DEPT     = 4
IDX_P1NAME   = 5
IDX_P1EMAIL  = 6
IDX_P2NAME   = 7
IDX_P2EMAIL  = 8
IDX_PHONE    = 9
IDX_REG_TIME = 10
IDX_STATUS   = 11
IDX_CHKTIME  = 12

EVENT_CODES = {
    'Pitch The Deck': 'PTD',
    'Plot Perfect':   'PPF',
    'One Piece':      'OPC',
    'Thinkathon':     'THK',
    'Pitch Perfect':  'PPR'
}
EVENTS = list(EVENT_CODES.keys())

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dv26_mcc_secret_2026_xK9pQ')

# ── Detect LAN IP for fallback check-in URLs ───────────────
def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return 'localhost'

LOCAL_IP = get_local_ip()

def get_checkin_base():
    env_base = os.environ.get('APP_URL', '').strip()
    if env_base:
        if 'datavista26.onrender.com' in env_base:
            env_base = env_base.replace('datavista26.onrender.com', 'datavista-26.onrender.com')
        return env_base.rstrip('/')
    try:
        if request and hasattr(request, 'host_url'):
            host = request.host_url.rstrip('/')
            if 'datavista26.onrender.com' in host:
                host = host.replace('datavista26.onrender.com', 'datavista-26.onrender.com')
            return host
    except Exception:
        pass
    return 'https://datavista-26.onrender.com'

# ── Mail config ───────────────────────────────────────
try:
    import mail_config as mc
    MAIL_OK = mc.MAIL_ENABLED and mc.MAIL_USERNAME != 'your.email@gmail.com'
except Exception:
    MAIL_OK = False

# ── Registration ID ───────────────────────────────────
def gen_reg_id(event):
    if event == 'Treasure Hunt': event = 'One Piece'
    code   = EVENT_CODES.get(event, 'OPC' if 'Piece' in str(event) else 'EVT')
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f'DV26-{code}-{suffix}'

# ── Excel helpers ─────────────────────────────────────
def init_excel():
    if not os.path.exists(EXCEL_PATH):
        _create_excel()
        return
    # Migrate: add missing columns
    wb  = openpyxl.load_workbook(EXCEL_PATH)
    ws  = wb.active
    existing = [c.value for c in ws[1]]
    changed  = False

    hf    = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    hfill = PatternFill('solid', fgColor='6B0F2A')
    ha    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin  = Side(style='thin', color='D4A843')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_col(header, width, default=''):
        nonlocal changed
        col = ws.max_column + 1
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = bdr
        ws.column_dimensions[get_column_letter(col)].width = width
        for r in range(2, ws.max_row + 1):
            if any(ws.cell(r, c).value for c in range(1, col)):
                ws.cell(r, col).value = default
        changed = True
        print(f'  [EXCEL] Migrated: added "{header}" column.')

    if 'Reg. ID' not in existing:
        ws.insert_cols(2)
        cell = ws.cell(row=1, column=2, value='Reg. ID')
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = bdr
        ws.column_dimensions['B'].width = 18
        for r in range(2, ws.max_row + 1):
            if any(ws.cell(r, c).value for c in range(1, ws.max_column + 1)):
                ws.cell(r, 2).value = 'DV26-MIG-XXXXX'
        changed = True
        print('  [EXCEL] Migrated: added "Reg. ID" column.')
        existing = [c.value for c in ws[1]]  # refresh

    if 'Check-In Status' not in [c.value for c in ws[1]]:
        add_col('Check-In Status', 15, 'Pending')
    if 'Check-In Time' not in [c.value for c in ws[1]]:
        add_col('Check-In Time', 20, '')

    if changed:
        wb.save(EXCEL_PATH)

def _create_excel():
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Registrations'
    ws.append(HEADERS)
    hf    = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    hfill = PatternFill('solid', fgColor='6B0F2A')
    ha    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin  = Side(style='thin', color='D4A843')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, cell in enumerate(ws[1], 1):
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci - 1]
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'
    wb.save(EXCEL_PATH)
    print(f'  [EXCEL] Created: {EXCEL_PATH}')

def style_data_row(ws, row_num):
    fill   = PatternFill('solid', fgColor='FDF5E6' if row_num % 2 == 0 else 'FFFFFF')
    thin   = Side(style='thin', color='E0C080')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        cell.fill = fill; cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

def find_row_by_reg_id(ws, reg_id):
    """Return (row_index, row_values) for the given Reg.ID, or (None, None)."""
    target = str(reg_id or '').strip().upper()
    if not target:
        return None, None
    for row_idx in range(2, ws.max_row + 1):
        cell_val = str(ws.cell(row=row_idx, column=IDX_REG_ID + 1).value or '').strip().upper()
        if cell_val == target:
            vals = [ws.cell(row=row_idx, column=c).value
                    for c in range(1, ws.max_column + 1)]
            return row_idx, vals
    return None, None

# ── Leaderboard helpers ───────────────────────────────
def sync_registrations_to_lb(lb_data):
    """Sync enrolled participants from Supabase into leaderboard.json."""
    if not isinstance(lb_data, dict):
        lb_data = {'entries': lb_data if isinstance(lb_data, list) else [], 'next_id': 1, 'last_updated': ''}

    existing = {}
    for entry in lb_data.get('entries', []):
        if isinstance(entry, dict) and entry.get('reg_id'):
            existing[entry['reg_id']] = entry

    new_entries = []

    try:
        db_rows = db_get_registrations()
        if db_rows:
            for r in db_rows:
                reg_id  = str(r.get('reg_id', '')).strip()
                if not reg_id: continue
                status  = str(r.get('status', 'Pending')).strip()
                
                # ONLY include participants who have CHECKED IN
                if status != 'Checked In':
                    continue
                event   = str(r.get('event', '')).strip()
                college = str(r.get('college', '')).strip()
                p1      = str(r.get('p1name', '')).strip()
                p2      = str(r.get('p2name', '')).strip()
                reg_on  = str(r.get('created_at', ''))[:16].replace('T', ' ') if r.get('created_at') else ''
                team    = f"{p1} & {p2}" if p2 else p1

                db_score = 0
                try: db_score = int(r.get('score', 0) or 0)
                except Exception: db_score = 0

                if reg_id in existing:
                    existing[reg_id]['event']   = event
                    existing[reg_id]['college'] = college
                    existing[reg_id]['team']    = team
                    if db_score != 0 or 'score' not in existing[reg_id]:
                        existing[reg_id]['score'] = db_score
                    new_entries.append(existing[reg_id])
                else:
                    entry = {
                        'id':        lb_data.get('next_id', 1),
                        'reg_id':    reg_id,
                        'event':     event,
                        'college':   college,
                        'team':      team,
                        'score':     db_score,
                        'timestamp': reg_on or datetime.now().strftime('%d-%m-%Y %H:%M')
                    }
                    lb_data['next_id'] = lb_data.get('next_id', 1) + 1
                    new_entries.append(entry)

            for entry in lb_data.get('entries', []):
                if isinstance(entry, dict) and not entry.get('reg_id') and entry not in new_entries:
                    new_entries.append(entry)

            lb_data['entries'] = new_entries
    except Exception as e:
        print(f"  [LB SYNC Error]: {e}")

    return lb_data

def load_lb():
    if not os.path.exists(LEADERBOARD_PATH):
        data = {'entries': [], 'next_id': 1, 'last_updated': ''}
    else:
        try:
            with open(LEADERBOARD_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if not isinstance(data, dict):
                data = {'entries': data if isinstance(data, list) else [], 'next_id': 1, 'last_updated': ''}
        except Exception:
            data = {'entries': [], 'next_id': 1, 'last_updated': ''}

    data = sync_registrations_to_lb(data)
    save_lb(data)
    return data

def save_lb(data):
    if not isinstance(data, dict):
        data = {'entries': data if isinstance(data, list) else [], 'next_id': 1, 'last_updated': ''}
    data['last_updated'] = datetime.now().strftime('%d-%m-%Y %H:%M')
    with open(LEADERBOARD_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

# ── Email helpers ─────────────────────────────────────
def build_confirmation_email(d, reg_id, checkin_url):

    def row(label, value):
        return f'''
        <tr>
          <td style="padding:10px 16px;border-bottom:1px solid #2A1020;
                     color:#9A7420;font-size:11px;font-weight:bold;
                     text-transform:uppercase;letter-spacing:1px;width:38%;">{label}</td>
          <td style="padding:10px 16px;border-bottom:1px solid #2A1020;
                     color:#F5ECD7;font-size:13px;">{value}</td>
        </tr>'''

    details_rows = (
        row('Event',         d['event']) +
        row('College',       d['college']) +
        row('Department',    d['department']) +
        row('Participant 1', d['p1name']) +
        row('Participant 2', d['p2name']) +
        row('Phone',         d['phone']) +
        row('Registered On', datetime.now().strftime('%d %B %Y, %H:%M'))
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Registration Confirmed — DATA VISTA '26</title></head>
<body style="margin:0;padding:0;background:#100105;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#100105;padding:30px 12px;">
<tr><td align="center">
<table width="100%" cellpadding="0" cellspacing="0"
       style="max-width:580px;background:#120208;border-radius:18px;overflow:hidden;border:1px solid #3A1525;">

  <tr>
    <td style="background:linear-gradient(135deg,#6B0F2A 0%,#3D0818 100%);
               padding:36px 30px 28px;text-align:center;border-bottom:2px solid #D4A843;">
      <p style="color:#C9A45A;font-size:10px;letter-spacing:5px;text-transform:uppercase;margin:0 0 6px;">
        Department of Data Science &bull; Madras Christian College</p>
      <h1 style="color:#F0C96A;font-size:28px;margin:0;letter-spacing:4px;
                 font-family:Georgia,serif;font-weight:bold;">
        DATA VISTA <span style="color:#D4A843;">'26</span></h1>
      <p style="color:#8A6A30;font-size:11px;margin:8px 0 0;letter-spacing:1px;">INTERCOLLEGIATE TECH FEST</p>
    </td>
  </tr>

  <tr>
    <td style="padding:32px 30px 20px;text-align:center;">
      <div style="width:64px;height:64px;background:linear-gradient(135deg,#D4A843,#9A7420);
                  border-radius:50%;margin:0 auto 16px;line-height:64px;font-size:26px;
                  color:#1a0205;font-weight:bold;">&#10003;</div>
      <h2 style="color:#D4A843;font-size:19px;margin:0 0 6px;letter-spacing:1px;">Registration Confirmed!</h2>
      <p style="color:#8A7050;font-size:13px;margin:0;line-height:1.6;">
        Your participation has been successfully registered.<br>
        We look forward to seeing you on <strong style="color:#C9A45A;">31st August 2026</strong>!</p>
    </td>
  </tr>

  <tr>
    <td style="padding:0 30px 24px;">
      <div style="background:#1E040D;border:2px solid #D4A843;border-radius:12px;padding:20px;text-align:center;">
        <p style="color:#6A5025;font-size:10px;text-transform:uppercase;letter-spacing:3px;margin:0 0 10px;">Your Registration ID</p>
        <h2 style="color:#F0C96A;font-size:21px;font-family:'Courier New',Courier,monospace;
                   letter-spacing:5px;margin:0;padding:6px 0;">{reg_id}</h2>
        <p style="color:#5A4020;font-size:11px;margin:8px 0 0;">&#128204;&nbsp; Keep this safe — present it at check-in</p>
      </div>
    </td>
  </tr>

  <tr>
    <td style="padding:0 30px 24px;text-align:center;">
      <p style="color:#6A5025;font-size:10px;text-transform:uppercase;letter-spacing:2px;margin:0 0 10px;">
        &#127919;&nbsp; Show this at the Registration Desk</p>
      <div style="background:#1E040D;border:2px dashed #D4A843;border-radius:12px;padding:18px 24px;display:inline-block;">
        <p style="color:#6A5025;font-size:9px;text-transform:uppercase;letter-spacing:3px;margin:0 0 8px;">Your Attendance ID</p>
        <p style="color:#F0C96A;font-size:26px;font-family:'Courier New',Courier,monospace;
                  letter-spacing:6px;font-weight:bold;margin:0;">{reg_id}</p>
        <p style="color:#4A3015;font-size:10px;margin:10px 0 0;">
          &#128205;&nbsp; Present this ID to the registration team on event day
        </p>
      </div>
    </td>
  </tr>

  <tr>
    <td style="padding:0 30px 24px;">
      <p style="color:#6A5025;font-size:10px;text-transform:uppercase;letter-spacing:2px;margin:0 0 10px;">Registration Details</p>
      <table width="100%" cellpadding="0" cellspacing="0"
             style="background:#1E040D;border-radius:10px;border:1px solid #2A1020;overflow:hidden;">
        {details_rows}
      </table>
    </td>
  </tr>

  <tr>
    <td style="background:#1A0310;padding:20px 30px;border-top:1px solid #2A1020;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr><td style="padding:4px 0;color:#9A7840;font-size:12px;">
          &#128197;&nbsp;<strong style="color:#D4A843;">Date:</strong>&nbsp; 31st August 2026</td></tr>
        <tr><td style="padding:4px 0;color:#9A7840;font-size:12px;">
          &#128205;&nbsp;<strong style="color:#D4A843;">Venue:</strong>&nbsp; Anderson Hall, Madras Christian College, Chennai</td></tr>
        <tr><td style="padding:4px 0;color:#9A7840;font-size:12px;">
          &#9200;&nbsp;<strong style="color:#D4A843;">Report By:</strong>&nbsp; 9:00 AM</td></tr>
      </table>
    </td>
  </tr>

  <tr>
    <td style="padding:20px 30px;background:#150208;border-top:1px solid #200A14;">
      <p style="color:#6A5025;font-size:10px;text-transform:uppercase;letter-spacing:2px;margin:0 0 10px;">Important Reminders</p>
      <p style="color:#7A6035;font-size:12px;margin:0;line-height:1.9;">
        &#x2022;&nbsp; Only <strong style="color:#C9A45A;">1 team per department</strong> is allowed per event.<br>
        &#x2022;&nbsp; Bring your valid <strong style="color:#C9A45A;">College ID card</strong> on the event day.<br>
        &#x2022;&nbsp; Show this email's <strong style="color:#C9A45A;">Registration ID</strong> at check-in for attendance marking.<br>
        &#x2022;&nbsp; Adhere to all event rules as per the official rule book.
      </p>
    </td>
  </tr>

  <tr>
    <td style="background:#0E0206;padding:18px 30px;text-align:center;border-top:1px solid #1A0510;">
      <p style="color:#4A3020;font-size:11px;margin:0 0 4px;">
        Organized by <strong style="color:#6A4A25;">DATONS</strong> — Department of Data Science Association</p>
      <p style="color:#3A2015;font-size:10px;margin:0;">Madras Christian College (Autonomous), Chennai</p>
    </td>
  </tr>

</table></td></tr></table>
</body></html>"""


def send_confirmation_async(to_emails, d, reg_id):
    def _worker():
        try:
            url = 'https://api.emailjs.com/api/v1.0/email/send'
            headers = {
                'Content-Type': 'application/json',
                'Origin': 'https://datavista-26.onrender.com',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
            }
            for email in to_emails:
                email_clean = email.strip()
                if not email_clean: continue
                payload = {
                    'service_id': 'service_ras66em',
                    'template_id': 'template_00s7x05',
                    'user_id': 'jPaO__3oPI2sybhNY',
                    'template_params': {
                        'to_email': email_clean,
                        'email': email_clean,
                        'user_email': email_clean,
                        'p1email': d.get('p1email', ''),
                        'p2email': d.get('p2email', ''),
                        'p1name': d.get('p1name', ''),
                        'p2name': d.get('p2name', ''),
                        'reg_id': reg_id,
                        'event': d.get('event', ''),
                        'college': d.get('college', ''),
                        'department': d.get('department', ''),
                        'phone': d.get('phone', ''),
                        'message': f"Registration ID: {reg_id}\nEvent: {d.get('event','')}\nCollege: {d.get('college','')}"
                    }
                }
                req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), headers=headers)
                with urllib.request.urlopen(req, timeout=15) as resp:
                    print(f'  [EMAIL] EmailJS HTTPS sent to {email_clean} (Status: {resp.status})')
        except Exception as err:
            import traceback
            print(f'  [EMAIL] EmailJS worker error: {err}')
            traceback.print_exc()

    t = threading.Thread(target=_worker)
    t.daemon = False
    t.start()


ADMIN_SECRET_TOKEN = 'DV26_ADMIN_SECURE_TOKEN_2026'

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        return f(*args, **kwargs)
    return decorated

# ── Supabase Persistent Cloud Database Config ──────────
SUPABASE_URL = 'https://dkylhswboedbxttfpkzo.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRreWxoc3dib2VkYnh0dF...0.1ziUobZ4kYucdJqJ6oGva-CcCCnq7I9-bvreJ59HkjE'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRreWxoc3dib2VkYnh0dGZwa3pvIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODU3Njk3NjAsImV4cCI6MjEwMTM0NTc2MH0.1ziUobZ4kYucdJqJ6oGva-CcCCnq7I9-bvreJ59HkjE'

import urllib.request
import urllib.parse

def db_get_registrations():
    url = f"{SUPABASE_URL}/rest/v1/registrations?select=*&order=id.asc"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json'
    }
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print('  [DB Error] db_get_registrations:', e)
        return []

def db_insert_registration(payload):
    url = f"{SUPABASE_URL}/rest/v1/registrations"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print('  [DB Error] db_insert_registration:', e)
        return None

def db_update_checkin(reg_id, status_val, checkin_time_str):
    url = f"{SUPABASE_URL}/rest/v1/registrations?reg_id=eq.{urllib.parse.quote(reg_id)}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }
    payload = {'status': status_val, 'checkin_time': checkin_time_str}
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), headers=headers, method='PATCH')
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print('  [DB Error] db_update_checkin:', e)
        return None

def db_delete_registration(reg_id):
    url = f"{SUPABASE_URL}/rest/v1/registrations?reg_id=eq.{urllib.parse.quote(reg_id)}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json'
    }
    req = urllib.request.Request(url, headers=headers, method='DELETE')
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return True
    except Exception as e:
        print('  [DB Error] db_delete_registration:', e)
        return False

def db_update_score(reg_id, score_val):
    if not reg_id: return None
    url = f"{SUPABASE_URL}/rest/v1/registrations?reg_id=eq.{urllib.parse.quote(reg_id)}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }
    payload = {'score': int(score_val)}
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), headers=headers, method='PATCH')
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print('  [DB Error] db_update_score:', e)
        return None

# ── Page routes (no-cache and CORS headers) ──────
from flask import make_response

@app.after_request
def add_cors_and_cache_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma']        = 'no-cache'
    response.headers['Expires']       = '0'
    return response

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/api/health')
def health_check():
    return jsonify({'status': 'ok', 'message': 'DataVista server is active'}), 200

@app.route('/register')
@app.route('/register.html')
def register_page():
    return send_from_directory(BASE_DIR, 'register.html')

@app.route('/admin')
@app.route('/admin.html')
def admin_page():
    return send_from_directory(BASE_DIR, 'admin.html')

@app.route('/leaderboard')
@app.route('/leaderboard.html')
def leaderboard_page():
    return send_from_directory(BASE_DIR, 'leaderboard.html')

@app.route('/checkin')
@app.route('/checkin.html')
@app.route('/checkin/<reg_id>')
def checkin_page(reg_id=None):
    return send_from_directory(BASE_DIR, 'checkin.html')

@app.route('/<path:filename>')
def static_files(filename):
    return send_from_directory(BASE_DIR, filename)

# ── Registration API ──────────────────────────────────
@app.route('/api/register', methods=['POST'])
def api_register():
    try:
        d = request.get_json(force=True, silent=True) or {}
        required = ['event','college','department','p1name','phone']
        for f in required:
            if not str(d.get(f,'') or '').strip():
                return jsonify({'success':False,'error':f'Missing: {f}'}), 400

        event_name = str(d.get('event','')).strip()
        college    = str(d.get('college','')).strip()
        dept       = str(d.get('department','')).strip()
        p1name     = str(d.get('p1name','')).strip()
        p1email    = str(d.get('p1email','') or '').strip()
        p2name     = str(d.get('p2name','') or '').strip()
        p2email    = str(d.get('p2email','') or '').strip()
        phone      = str(d.get('phone','')).strip()
        
        is_spot    = d.get('is_spot') or (p1email and 'spot@' in p1email)
        status_val = 'Checked In' if is_spot else 'Pending'
        chk_time   = datetime.now().strftime('%d-%m-%Y %H:%M:%S') if is_spot else ''

        reg_id      = gen_reg_id(event_name)
        checkin_url = f'{get_checkin_base()}/checkin/{reg_id}'

        # Insert into Supabase Cloud Database (Permanent)
        db_insert_registration({
            'reg_id': reg_id,
            'event': event_name,
            'college': college,
            'department': dept,
            'p1name': p1name,
            'p1email': p1email,
            'p2name': p2name,
            'p2email': p2email,
            'phone': phone,
            'status': status_val,
            'checkin_time': chk_time
        })

        try:
            wb  = openpyxl.load_workbook(EXCEL_PATH)
            ws  = wb.active
            sno = ws.max_row
            ws.append([
                sno, reg_id,
                event_name, college, dept,
                p1name, p1email,
                p2name, p2email,
                phone,
                datetime.now().strftime('%d-%m-%Y %H:%M'),
                status_val, chk_time
            ])
            style_data_row(ws, ws.max_row)
            wb.save(EXCEL_PATH)
        except Exception as ex: print('Excel save warn:', ex)

        # Auto-sync Leaderboard
        load_lb()

        import re
        email_pattern = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
        raw_emails = [p1email, p2email]
        to_list = list(set([e for e in raw_emails if e and email_pattern.match(e)]))

        email_sent = False
        if to_list:
            send_confirmation_async(to_list, {
                'event': event_name, 'college': college, 'department': dept,
                'p1name': p1name, 'p1email': p1email, 'p2name': p2name, 'p2email': p2email,
                'phone': phone
            }, reg_id)
            email_sent = True

        return jsonify({'success':True, 'reg_id':reg_id, 'email_sent':email_sent})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)}), 500

# ── Public Check-In API ───────────────────────────────
@app.route('/api/checkin/<reg_id>')
def api_checkin_get(reg_id):
    """Return registration details for the check-in page."""
    db_rows = db_get_registrations()
    match = next((r for r in db_rows if r.get('reg_id') == reg_id), None)
    if match:
        return jsonify({
            'found':      True,
            'reg_id':     reg_id,
            'event':      match.get('event', ''),
            'college':    match.get('college', ''),
            'department': match.get('department', ''),
            'p1name':     match.get('p1name', ''),
            'p2name':     match.get('p2name', ''),
            'phone':      match.get('phone', ''),
            'status':     match.get('status', 'Pending'),
            'checkin_at': match.get('checkin_time', '')
        })

    # Fallback to local Excel file
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    row_idx, vals = find_row_by_reg_id(ws, reg_id)
    if row_idx is None:
        return jsonify({'found': False, 'error': 'Registration ID not found. Please check and try again.'}), 200

    headers = [c.value for c in ws[1]]
    def get(col_name):
        try:
            idx = headers.index(col_name)
            v   = vals[idx]
            return str(v) if v is not None else ''
        except ValueError:
            return ''

    return jsonify({
        'found':      True,
        'reg_id':     reg_id,
        'event':      get('Event'),
        'college':    get('College'),
        'department': get('Department'),
        'p1name':     get('Participant 1 Name'),
        'p2name':     get('Participant 2 Name'),
        'phone':      get('Phone'),
        'status':     get('Check-In Status') or 'Pending',
        'checkin_at': get('Check-In Time')
    })

@app.route('/api/checkin/<reg_id>/mark', methods=['POST'])
def api_checkin_mark(reg_id):
    """Mark a participant as checked in."""
    now_str = datetime.now().strftime('%d-%m-%Y %H:%M:%S')

    # Update Supabase Cloud DB
    db_update_checkin(reg_id, 'Checked In', now_str)

    # Local Excel Fallback
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        row_idx, vals = find_row_by_reg_id(ws, reg_id)
        if row_idx is not None:
            headers  = [c.value for c in ws[1]]
            stat_col = headers.index('Check-In Status') + 1
            time_col = headers.index('Check-In Time')   + 1
            ws.cell(row=row_idx, column=stat_col).value = 'Checked In'
            ws.cell(row=row_idx, column=time_col).value  = now_str
            green = PatternFill('solid', fgColor='D4EDDA')
            ws.cell(row=row_idx, column=stat_col).fill = green
            wb.save(EXCEL_PATH)
    except Exception as ex:
        print('Excel checkin warn:', ex)

    print(f'  [CHECKIN] {reg_id} marked present at {now_str}')
    return jsonify({'success':True, 'checked_in_at': now_str})

# ── Public leaderboard API ────────────────────────────
@app.route('/api/leaderboard')
def api_leaderboard_public():
    data = load_lb()
    return jsonify({'entries':data.get('entries',[]),
                    'last_updated':data.get('last_updated','')})

# ── Admin login / logout ──────────────────────────────
@app.route('/api/admin/login', methods=['POST', 'OPTIONS'])
def api_login():
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    d = request.get_json(force=True, silent=True) or {}
    user = str(d.get('username', '') or '').strip().lower()
    pwd  = str(d.get('password', '') or '').strip()

    valid_users = ['admin', ADMIN_USER.lower()]
    valid_passwords = [
        ADMIN_PASS,
        'datavista@26', 'datavista26', 'DataVista@26', 'DataVista26',
        'admin', 'admin123'
    ]

    if user in valid_users and (pwd in valid_passwords or pwd.lower() == ADMIN_PASS.lower()):
        session['admin_logged_in'] = True
        return jsonify({'success': True, 'token': ADMIN_SECRET_TOKEN})

    return jsonify({'success': False, 'message': 'Invalid credentials. Use Username: admin | Password: DataVista@26'}), 401

@app.route('/api/admin/logout', methods=['POST', 'OPTIONS'])
def api_logout():
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    session.clear()
    return jsonify({'success':True})

# ── Admin registrations ───────────────────────────────
@app.route('/api/admin/registrations')
@admin_required
def api_registrations():
    headers = ['S.No', 'Reg. ID', 'Event', 'College', 'Department', 'Participant 1 Name', 'Participant 1 Email', 'Participant 2 Name', 'Participant 2 Email', 'Phone', 'Registered On', 'Check-In Status', 'Check-In Time']
    rows, counts, checkin_counts = [], {}, {}

    db_rows = db_get_registrations()
    if db_rows:
        for idx, r in enumerate(db_rows, start=1):
            created_str = r.get('created_at', '')
            if created_str and 'T' in created_str:
                created_str = created_str[:16].replace('T', ' ')
            rl = [
                str(idx),
                r.get('reg_id', ''),
                r.get('event', ''),
                r.get('college', ''),
                r.get('department', ''),
                r.get('p1name', ''),
                r.get('p1email', ''),
                r.get('p2name', ''),
                r.get('p2email', ''),
                r.get('phone', ''),
                created_str,
                r.get('status', 'Pending'),
                r.get('checkin_time', '')
            ]
            rows.append(rl)
            ev = r.get('event', '')
            if ev == 'Treasure Hunt': ev = 'One Piece'
            st = r.get('status', 'Pending')
            counts[ev] = counts.get(ev, 0) + 1
            if st == 'Checked In':
                checkin_counts[ev] = checkin_counts.get(ev, 0) + 1
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rl = [str(v) if v is not None else '' for v in row]
                rows.append(rl)
                ev  = rl[IDX_EVENT]   if len(rl) > IDX_EVENT   else ''
                st  = rl[IDX_STATUS]  if len(rl) > IDX_STATUS  else 'Pending'
                counts[ev]        = counts.get(ev, 0) + 1
                if st == 'Checked In':
                    checkin_counts[ev] = checkin_counts.get(ev, 0) + 1

    return jsonify({'headers':headers,'rows':rows,'total':len(rows),
                    'counts':counts,'checkin_counts':checkin_counts})

@app.route('/api/admin/download', methods=['GET', 'OPTIONS'])
@admin_required
def api_download():
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    try:
        import io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Registrations'
        headers = ['S.No', 'Reg. ID', 'Event', 'College', 'Department', 'Participant 1 Name', 'Participant 1 Email', 'Participant 2 Name', 'Participant 2 Email', 'Phone', 'Registered On', 'Check-In Status', 'Check-In Time']
        ws.append(headers)
        
        db_rows = db_get_registrations()
        if db_rows:
            for idx, r in enumerate(db_rows, start=1):
                created_str = str(r.get('created_at', '') or '')
                if created_str and 'T' in created_str:
                    created_str = created_str[:16].replace('T', ' ')
                ws.append([
                    idx,
                    str(r.get('reg_id','') or ''),
                    str(r.get('event','') or ''),
                    str(r.get('college','') or ''),
                    str(r.get('department','') or ''),
                    str(r.get('p1name','') or ''),
                    str(r.get('p1email','') or ''),
                    str(r.get('p2name','') or ''),
                    str(r.get('p2email','') or ''),
                    str(r.get('phone','') or ''),
                    created_str,
                    str(r.get('status','Pending') or 'Pending'),
                    str(r.get('checkin_time','') or '')
                ])
                try: style_data_row(ws, ws.max_row)
                except Exception: pass
        
        buf = io.BytesIO()
        wb.save(buf)
        excel_data = buf.getvalue()

        response = make_response(excel_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=DataVista26_Registrations.xlsx'
        return response
    except Exception as e:
        import traceback
        print('  [DOWNLOAD ERROR]:', e)
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/registrations/<reg_id>', methods=['DELETE', 'OPTIONS'])
@admin_required
def api_delete_registration(reg_id):
    """Delete a registration by Reg. ID from Supabase and Excel."""
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    try:
        # Delete from Supabase Cloud Database
        db_delete_registration(reg_id)

        # Delete from local Excel file
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            row_idx, vals = find_row_by_reg_id(ws, reg_id)
            if row_idx is not None:
                ws.delete_rows(row_idx)
                for r in range(2, ws.max_row + 1):
                    if any(ws.cell(r, c).value for c in range(2, ws.max_column + 1)):
                        ws.cell(r, 1).value = r - 1
                wb.save(EXCEL_PATH)
        except Exception as ex:
            print('Excel delete warn:', ex)

        # Auto-prune from Leaderboard
        load_lb()
        print(f'  [ADMIN] Deleted registration: {reg_id}')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ── Admin leaderboard CRUD ────────────────────────────
@app.route('/api/admin/leaderboard', methods=['GET'])
@admin_required
def api_lb_get():
    return jsonify(load_lb())

@app.route('/api/admin/leaderboard/add', methods=['POST', 'OPTIONS'])
@admin_required
def api_lb_add():
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    d    = request.get_json(force=True, silent=True) or {}
    data = load_lb()
    entry = {
        'id':        data['next_id'],
        'event':     d.get('event','').strip(),
        'college':   d.get('college','').strip(),
        'team':      d.get('team','').strip(),
        'score':     int(d.get('score',0)),
        'timestamp': datetime.now().strftime('%d-%m-%Y %H:%M')
    }
    if not entry['event'] or not entry['college']:
        return jsonify({'success':False,'error':'Event and College required'}), 400
    data['entries'].append(entry); data['next_id'] += 1
    save_lb(data)
    return jsonify({'success':True,'entry':entry})

@app.route('/api/admin/leaderboard/<entry_id>', methods=['PUT', 'OPTIONS'])
@admin_required
def api_lb_update(entry_id):
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    d = request.get_json(force=True, silent=True) or {}
    data = load_lb()
    target_str = str(entry_id).strip().upper()
    for entry in data['entries']:
        if str(entry.get('id')) == target_str or str(entry.get('reg_id','')).strip().upper() == target_str:
            if 'event' in d:   entry['event']   = d['event']
            if 'college' in d: entry['college'] = d['college']
            if 'team' in d:    entry['team']    = d['team']
            if 'score' in d:
                entry['score'] = int(d['score'])
                if entry.get('reg_id'):
                    db_update_score(entry['reg_id'], entry['score'])
            entry['timestamp'] = datetime.now().strftime('%d-%m-%Y %H:%M')
            save_lb(data)
            return jsonify({'success':True,'entry':entry})
    return jsonify({'success':False,'error':'Not found'}), 404

@app.route('/api/admin/leaderboard/<int:entry_id>', methods=['DELETE', 'OPTIONS'])
@admin_required
def api_lb_delete(entry_id):
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    data   = load_lb()
    before = len(data['entries'])
    data['entries'] = [e for e in data['entries'] if e['id'] != entry_id]
    if len(data['entries']) == before:
        return jsonify({'success':False,'error':'Not found'}), 404
    save_lb(data)
    return jsonify({'success':True})

# ── Run ───────────────────────────────────────────────
if __name__ == '__main__':
    init_excel()
    mail_status = 'ENABLED' if MAIL_OK else 'DISABLED (fill mail_config.py)'
    print('\n' + '=' * 57)
    print("  DATA VISTA '26 - Server Running")
    print('=' * 57)
    print(f'  Main Site    -> http://{LOCAL_IP}:5000')
    print(f'  Leaderboard  -> http://{LOCAL_IP}:5000/leaderboard')
    print(f'  Admin        -> http://{LOCAL_IP}:5000/admin')
    print(f'  Check-In URL -> http://{LOCAL_IP}:5000/checkin/<REG_ID>')
    print('  Username     -> ' + ADMIN_USER)
    print('  Password     -> ' + ADMIN_PASS)
    print('  Email        -> ' + mail_status)
    print('=' * 57 + '\n')
    import sys
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    app.run(debug=False, port=5000, host='0.0.0.0')
